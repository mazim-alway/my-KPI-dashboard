import openpyxl
from db.queries import (
    add_project, add_period, add_theme,
)
from datetime import datetime


# ==========================================================
# Helper: Safe value extractor
# ==========================================================
def safe(cell):
    return cell.value if cell.value is not None else ""


# ==========================================================
# Helper: Convert Excel date to Python date
# ==========================================================
def parse_excel_date(value):
    if isinstance(value, datetime):
        return value.date()
    try:
        return datetime.strptime(str(value), "%m/%d/%Y").date()
    except:
        return None


# ==========================================================
# Extract period code from sheet name
# ==========================================================
def extract_period(sheet_title: str):
    """
    Example:
        '24F2 KPI Monitoring' → '24F2'
        '25F1' → '25F1'
    """
    parts = sheet_title.split()
    for p in parts:
        if "F" in p:
            return p.strip()
    return sheet_title  # fallback


# ==========================================================
# Mapping Excel sheet → database record
# ==========================================================
def process_sheet(sheet):
    """
    Generic parser for project rows from your XLSM template.
    Columns vary slightly but are normalized here.
    """

    period_code = extract_period(sheet.title)
    add_period(period_code)  # ensures period exists

    # Default header markers
    header_row = None

    # Find header row automatically
    for i in range(1, 10):
        row_vals = [safe(c) for c in sheet[i]]
        if "Project" in row_vals:
            header_row = i
            break

    if not header_row:
        print(f"⚠ No valid header row in sheet {sheet.title}")
        return

    # Column index mapping
    header = [safe(c) for c in sheet[header_row]]
    col_map = {name: index for index, name in enumerate(header)}

    # Iterate project rows below header
    for r in range(header_row + 1, sheet.max_row + 1):
        row = sheet[r]
        if not safe(row[col_map["Project"]]):
            continue  # skip blank rows

        # Extract fields
        theme = safe(row[col_map.get("Theme", 0)])
        project_name = safe(row[col_map["Project"]])
        product_item = safe(row[col_map.get("Product Item", 0)])
        process_name = safe(row[col_map.get("Process", 0)])
        details = safe(row[col_map.get("Details", 0)])
        remark = safe(row[col_map.get("Remark", 0)])
        deadline = parse_excel_date(row[col_map.get("Deadline", col_map.get("Dateline", 0))].value)

        registered_by = safe(row[col_map.get("Registered by", 0)])
        registered_on = safe(row[col_map.get("Registered on", 0)])
        status = safe(row[col_map.get("Status", 0)])
        kpi_value = safe(row[col_map.get("KPI", 0)])
        details = safe(row[col_map.get("Details", 0)])

        # Ensure theme exists
        if theme:
            add_theme(theme)

        # Insert into DB
        add_project(
            theme_id=None,              # will resolve theme later in future upgrade
            period_id=None,             # will resolve period ID in GUI part
            project_name=project_name,
            product_item=product_item,
            process_name=process_name,
            details=details,
            deadline=deadline,
            remark=remark,
            registered_by=registered_by,
            registered_on=registered_on,
            status=status,
            kpi_value=float(kpi_value) if str(kpi_value).replace('.', '', 1).isdigit() else None
        )

    print(f"✔ Imported sheet: {sheet.title}")


# ==========================================================
# Master Import Function
# ==========================================================
def import_xlsm(filepath: str):
    print(f"📥 Importing XLSM: {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Your file uses:
    # - Sheet1
    # - Sheet3
    # - 25F1
    # - Some sheets not relevant
    target_sheets = ["Sheet1", "Sheet3", "25F1"]

    for sheet_name in target_sheets:
        if sheet_name in wb.sheetnames:
            process_sheet(wb[sheet_name])
        else:
            print(f"⚠ Sheet missing: {sheet_name}")

    print("🎉 Import Completed Successfully!")
